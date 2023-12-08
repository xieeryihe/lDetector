import os
import threading
import time
import warnings

import openai
import openpyxl
import requests
from tenacity import retry, RetryError, wait_random, stop_after_attempt

warnings.filterwarnings("ignore")

RED_BG = "\033[41m"
GREEN_BG = "\033[42m"
END = "\033[0m"

GPT_MODEL = "gpt-3.5-turbo"
openai.api_key = "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"
# 检验
sys_template = '''
假设你是电商平台的搜索助理。该平台支持对各类商品和店铺的搜索，除搜索关键词外，返回结果时还考虑地理位置，优先返回附近的店铺和商品。
可以构建多种搜索关键词在电商平台上对店铺q进行搜索。搜索关键词的构造通常考虑店铺q的名称、店铺位置、店铺类型，店铺提供的服务等。
你的任务是已知店铺q和店铺类型q_type, 判断用户构建的搜索关键词keyword是否合理,合理的标准是能够完成对店铺q进行搜索，并且符合普通用户表达习惯。
judgement是你的判断结果，我将对你的判断结果进行解析，请严格按照要求输出，keyword合理输出“yes”,不合理输出“no”。
以下是几个合理的keyword，可供参考。
正例一 q="那时新疆·车师（大宁音乐广场店）"，q_type='新疆菜' ，keyword='大宁 新疆菜'，judgement='yes'
正例二 q="涟泉小町·日式SPA（梅陇店）"，q_type='按摩/足疗'， keyword='做SPA'，judgement='yes'
正例三 q="蕃茄田艺术（宝山高境中心）"，q_type='美术'， keyword='画画 附近'，judgement='yes'
以下是几个不合理的keyword,可供参考。
反例一 q="小月半美甲美睫"，q_type='美甲' ，keyword='睫 小月半美甲美'，judgement='no'
这个keyword不合理的原因是“睫”不是可以售卖的商品或服务，因为断句错误而不符合普通用户表达习惯
反例二 q="澜斐儿（田图苑店）"，q_type='美甲'， keyword='美甲公众号'，judgement='no'
这个keyword不合理的原因是“美甲公众号”不是可以售卖的商品或服务，不符合普通用户表达习惯，并且不能对店铺q"澜斐儿（田图苑店）"完成搜索
反例三 q="东尚造型（永清路店）"，q_type='美发'， keyword='东尚附近剪头发'，judgement='no'
这个例子错误的原因是“东尚”是店名，不是地名，因此“东尚附近”不符合普通用户表达习惯，不是合理的搜索关键词。
反例四 q="海底捞（五角场合生汇店）"，q_type='火锅'， keyword='合生汇'，judgement='no'
这个例子错误的原因是“合生汇”是一个商场，范围远大于店铺q"海底捞（五角场合生汇店）"，不能完成对"海底捞（五角场合生汇店）"的搜索，因此不是合理的搜索关键词
'''

results = {}
shop_name_col = 1
shop_type_col = 5
keyword_col = 6
judgement_col = 7
origin_line_num = 2
target_line_num = 2


class MyThread(threading.Thread):
    def __init__(self, target, row_num, args=()):
        super().__init__()
        self.target = target
        self.row_num = row_num
        self.args = args
        self._result = None

    def run(self):
        # print(f"now process with args {self.args}")
        self._result = self.target(*self.args)
        print(f"row: {self.row_num} complete, args={self.args}, judgement={self._result}")

    def get_result(self):
        return self._result


def write_results(origin_excel, target_excel):
    global origin_line_num, target_line_num
    origin_workbook = openpyxl.load_workbook(origin_excel)
    target_workbook = openpyxl.load_workbook(target_excel)
    origin_sheet = origin_workbook['Sheet1']
    target_sheet = target_workbook['Sheet1']
    origin_col_count = origin_sheet.max_column
    origin_row_count = origin_sheet.max_row
    # print("origin row count", origin_row_count)
    for i in range(2, origin_row_count + 1):
        judgement = results[i]
        # copy origin info
        for col in range(1, origin_col_count + 1):
            copy_cell = origin_sheet.cell(row=i, column=col).value
            target_sheet.cell(row=target_line_num, column=col, value=copy_cell)
        # write judgement
        target_sheet.cell(row=target_line_num, column=judgement_col, value=judgement)
        origin_line_num += 1
        target_line_num += 1

    target_workbook.save(target_excel)


@retry(wait=wait_random(min=1, max=5), stop=stop_after_attempt(3))
def chat_completion_request(messages, model=GPT_MODEL):
    print(f"{GREEN_BG} Info {END}, Generating ChatCompletion response")
    headers = {
        "Content-Type": "application/json",
        "Authorization": "Bearer " + openai.api_key,
    }
    json_data = {"model": model, "temperature": 0, "messages": messages}
    try:
        response = requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers=headers,
            json=json_data,
            timeout=30,
        )
        return response
    except Exception as e:
        print(f"{RED_BG} Warning {END}, Unable to generate ChatCompletion response")
        print(f"Exception: {type(e)} {e}")
        raise e


def get_judgement(q, q_type, keyword):
    messages = []
    human_template = f"现在q={q},q_type={q_type},keyword={keyword},judgement="
    input_template = sys_template + human_template
    messages.append({"role": "user", "content": input_template})
    try:
        chat_response = chat_completion_request(messages)
        # print(chat_response.json())
        assistant_message = chat_response.json()["choices"][0]["message"]
        content = assistant_message['content']
    except RetryError:
        print("RetryError")
        content = "RetryError"
    except (AttributeError, KeyError) as e:
        content = "json error"
        print(e)
    # print(content)
    return content


def process(excel_name):
    workbook = openpyxl.load_workbook(excel_name)
    sheet = workbook['Sheet1']
    row_count = sheet.max_row
    threads = []
    for row_num in range(2, row_count + 1):
        q = sheet.cell(row=row_num, column=shop_name_col).value
        if q is None:
            continue
        q_type = sheet.cell(row=row_num, column=shop_type_col).value
        keyword = sheet.cell(row=row_num, column=keyword_col).value
        my_thread = MyThread(target=get_judgement, row_num=row_num, args=(q, q_type, keyword))
        threads.append(my_thread)
    i = 0
    for t in threads:
        t.start()
        i += 1
        if i % 10 == 0 and i != 0:
            time.sleep(5)
        # gpt-3.5-turbo tokens per min (TPM) = 90000
        # https://platform.openai.com/account/limits
    for t in threads:
        t.join()
    for t in threads:
        judgement = t.get_result()
        # print(f"row: {t.row_num},args: {t.args}, judgement: {judgement}")
        results[t.row_num] = judgement


def clean_excel(filename):
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active
    for row in worksheet.iter_rows():
        for cell in row:
            cell.value = None
    workbook.save(filename)


def main():
    global results
    origin_excel = "gpt_gen_zero_shot.xlsx"
    target_excel = "gpt_gen_zero_shot_check.xlsx"
    # if target_excel does not exist, then create
    if not os.path.exists(target_excel):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        workbook.save(target_excel)
        print("create file "+target_excel)
    # else exists, clean it.
    else:
        print(f"clean file: {target_excel}")
        clean_excel(target_excel)

    print(f" origin excel:{origin_excel}, \n output_excel:{target_excel}.")
    process(excel_name=origin_excel)
    write_results(origin_excel, target_excel)
    return


if __name__ == "__main__":
    main()
