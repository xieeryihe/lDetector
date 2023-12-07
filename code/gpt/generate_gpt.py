import re
import threading
import time
import os
import openai
import openpyxl
import requests
from tenacity import retry, wait_random, stop_after_attempt

RED_BG = "\033[41m"
GREEN_BG = "\033[42m"
END = "\033[0m"

GPT_MODEL = "gpt-3.5-turbo"
openai.api_key = "your key"
# 生成
sys_template_cot = '''
target_shop: 那时新疆·车师（大宁音乐广场店），q_type:'新疆菜' 
Q: 以json格式对一组能够在电商平台上对target_shop进行搜索的搜索关键词q_list进行输出。q_list=？
A: 电商平台支持对各类商品和店铺的搜索，返回搜索结果时还考虑地理位置，优先返回附近的店铺和商品。
首先，target_shop的名称是"那时新疆·车师（大宁音乐广场店）",搜索店铺全称可以搜到该店，又从店铺名称出发，可搜索"那时新疆"，"车师"。
其次，从q_type看，target_shop应该是一家新疆菜餐厅，从店铺类型出发，可搜索"新疆菜"，"新疆餐厅" 。
再次，新疆菜餐厅通常提供大盘鸡、抓饭等菜品。从代表性菜品出发，可搜索"大盘鸡"，"抓饭" 。
最后，target_shop位于大宁音乐广场，从地理位置出发，可搜索"新疆菜 大宁广场"，"车师 大宁音乐广场" ，"附近大盘鸡"。
因此：
q_list={{"keywords":[
"那时新疆·车师（大宁音乐广场店）",
"那时新疆",
"车师",
"新疆菜",
"新疆餐厅",
"大盘鸡",
"抓饭",
"新疆菜 大宁广场",
"车师 大宁音乐广场",
"附近大盘鸡"
]}}
target_shop: 涟泉小町·日式SPA（梅陇店），q_type:'按摩/足疗'
Q: 以json格式对一组能够在电商平台上对target_shop进行搜索的搜索关键词q_list进行输出。q_list=？
A: 电商平台支持对各类商品和店铺的搜索，返回搜索结果时还考虑地理位置，优先返回附近的店铺和商品。
首先，target_shop的名称是"涟泉小町·日式SPA（梅陇店）",搜索店铺全称可以搜到该店，又从店铺名称出发，可搜索"涟泉小町"，"日式SPA"。
其次，从q_type看，target_shop应该是一家按摩/足疗店，从店铺类型出发，可搜索"做SPA"，"做足疗"，"澡堂" 。
再次，按摩/足疗店通常提供泡澡、养生按摩等服务。从代表性服务出发，可搜索"泡澡"，"养生按摩" 。
最后，target_shop位于梅陇，从地理位置出发，可搜索"泡澡 梅陇"，"养生按摩 梅陇" ，"附近泡澡"。
因此：
q_list={{"keywords":[
"涟泉小町·日式SPA（梅陇店）",
"涟泉小町",
"日式SPA",
"做SPA",
"做足疗",
"澡堂",
"泡澡",
"养生按摩",
"泡澡 梅陇",
"养生按摩 梅陇",
"附近泡澡"
]}}
target_shop: 华润万家（三门路店），q_type:'超市'
Q: 以json格式对一组能够在电商平台上对target_shop进行搜索的搜索关键词q_list进行输出。q_list=？
A: 电商平台支持对各类商品和店铺的搜索，返回搜索结果时还考虑地理位置，优先返回附近的店铺和商品。
首先，target_shop的名称是"华润万家（三门路店）",搜索店铺全称可以搜到该店，又从店铺名称出发，可搜索"华润万家"，"华润"。
其次，从q_type看，target_shop应该是一家超市，从店铺类型出发，可搜索"大超市"，"华润超市"，"华润万家超市" ，"综合超市"，"逛超市"。
再次，超市通常水果生鲜等商品。从代表性商品出发，可搜索"超市买菜"。
最后，target_shop位于三门路，从地理位置出发，可搜索"超市 三门路"，"华润万家 三门路" ，"华润超市 三门路"，"附近华润万家"。
因此：
q_list={{"keywords":[
"华润万家（三门路店）",
"华润万家",
"华润",
"大超市",
"华润超市",
"华润万家超市",
"综合超市",
"超市买菜",
"超市 三门路",
"华润万家 三门路",
"华润超市 三门路",
"附近华润万家"
]}}
target_shop: 万达影城（五角场万达广场IMAX店）q_type:'电影院'
Q: 以json格式对一组能够在电商平台上对target_shop进行搜索的搜索关键词q_list进行输出。q_list=？
A: 电商平台支持对各类商品和店铺的搜索，返回搜索结果时还考虑地理位置，优先返回附近的店铺和商品。
首先，target_shop的名称是"万达影城（五角场万达广场IMAX店）",搜索店铺全称可以搜到该店，又从店铺名称出发，可搜索"万达影城"，"影城"。
其次，从q_type看，target_shop应该是一家电影院，从店铺类型出发，可搜索"电影院"，"影院"。
再次，电影院通常提供观影服务。从代表性服务出发，可搜索"看电影"，"IMAX电影"，"3D电影"，"电影新片"。
最后，target_shop位于五角场万达广场，从地理位置出发，可搜索"电影院 五角场"，"万达广场的电影院" ，"万达影城 五角场"，"附近的电影院"。
因此：
q_list={{"keywords":[
"万达影城（五角场万达广场IMAX店）",
"万达影城",
"影城",
"电影院",
"影院",
"看电影",
"IMAX电影",
"3D电影",
"电影新片",
"影院 五角场",
"万达广场的电影院",
"万达影城 五角场",
"附近的电影院"
]}}
target_shop: 阿沐汽车工作室，q_type:'汽车维修'
Q: 以json格式对一组能够在电商平台上对target_shop进行搜索的搜索关键词q_list进行输出。q_list=？
A: 电商平台支持对各类商品和店铺的搜索，返回搜索结果时还考虑地理位置，优先返回附近的店铺和商品。
首先，target_shop的名称是"阿沐汽车工作室",搜索店铺全称可以搜到该店，又从店铺名称出发，可搜索"阿沐汽车"，"汽车工作室"。
其次，从q_type看，target_shop应该是一家汽车维修店，从店铺类型出发，可搜索"修车铺"，"修车店"，"修汽车"。
再次，汽车维修店提供换轮胎，汽车喷漆，汽车改装等服务。从代表性服务出发，可搜索"汽车换轮胎"，"汽车喷漆"，"汽车改装"。
最后，target_shop未提供地理位置信息，从地理位置出发，可搜索"修车店 附近"，"附近修汽车的地方" 。
因此：
q_list={{"keywords":[
"阿沐汽车工作室",
"阿沐汽车",
"汽车工作室",
"修车铺",
"修车店",
"修汽车",
"汽车换轮胎",
"汽车喷漆",
"汽车改装",
"修车店 附近",
"附近修汽车的地方"
]}}
保证生成的 keywords 列表最多包含 7 个条目。
'''

sys_template_few_shot = '''
target_shop: 那时新疆·车师（大宁音乐广场店），q_type:'新疆菜' 
Q: 以json格式对一组能够在电商平台上对target_shop进行搜索的搜索关键词q_list进行输出。q_list=？
A:
q_list={{"keywords":[
"那时新疆·车师（大宁音乐广场店）",
"那时新疆",
"车师",
"新疆菜"
"新疆餐厅"
"大盘鸡"
"抓饭",
"新疆菜 大宁广场",
"车师 大宁音乐广场",
"附近大盘鸡"
]}}
target_shop: 涟泉小町·日式SPA（梅陇店），q_type:'按摩/足疗'
Q: 以json格式对一组能够在电商平台上对target_shop进行搜索的搜索关键词q_list进行输出。q_list=？
A:
q_list={{"keywords":[
"涟泉小町·日式SPA（梅陇店）",
"涟泉小町",
"日式SPA",
"做SPA"
"做足疗"
"澡堂"
"泡澡",
"养生按摩",
"泡澡 梅陇",
"养生按摩 梅陇",
"附近泡澡"
]}}
target_shop: 华润万家（三门路店），q_type:'超市'
Q: 以json格式对一组能够在电商平台上对target_shop进行搜索的搜索关键词q_list进行输出。q_list=？
A:
q_list={{"keywords":[
"华润万家（三门路店）",
"华润万家",
"华润",
"大超市"
"华润超市"
"华润万家超市"
"综合超市",
"超市买菜",
"超市 三门路",
"华润万家 三门路",
"华润超市 三门路",
"附近华润万家"
]}}
target_shop: 万达影城（五角场万达广场IMAX店）q_type:'电影院'
Q: 以json格式对一组能够在电商平台上对target_shop进行搜索的搜索关键词q_list进行输出。q_list=？
A:
q_list={{"keywords":[
"万达影城（五角场万达广场IMAX店）",
"万达影城",
"影城",
"电影院"
"影院"
"看电影"
"IMAX电影",
"3D电影",
"电影新片",
"影院 五角场",
"万达广场的电影院",
"万达影城 五角场",
"附近的电影院"
]}}
target_shop: 阿沐汽车工作室，q_type:'汽车维修'
Q: 以json格式对一组能够在电商平台上对target_shop进行搜索的搜索关键词q_list进行输出。q_list=？
A:
q_list={{"keywords":[
"阿沐汽车工作室",
"阿沐汽车",
"汽车工作室",
"修车铺"
"修车店"
"修汽车"
"汽车换轮胎",
"汽车喷漆",
"汽车改装",
"修车店 附近",
"附近修汽车的地方"
]}}
保证生成的 keywords 列表最多包含 7 个条目。
'''

sys_template_zero_shot = '''
假设你是电商平台的一个用户。该平台支持对各类商品和店铺的搜索，除搜索关键词外，返回结果时还考虑地理位置，优先返回附近的店铺和商品。
已知店铺 q 和店铺类型 q_type , 你可以构建多种搜索关键词在电商平台上对店铺q进行搜索。搜索关键词的构造通常考虑店铺的名称、店铺位置、店铺类型，店铺提供的服务等。
一组能在电商网站上达成相同搜索目的搜索关键词称为一组等价的搜索关键词q_list。
你的任务是构造出这一组等价搜索关键词q_list，并采用Json格式输出，保证生成的 keywords 列表最多包含 7 个条目。仿照下述格式。
q_list={{"keywords":[
item1,
item2,
item3,
item4,
item5,
item6,
item7
]}}
'''


results = {}
shop_name_col = 1
shop_type_col = 5
empty_col = 6
origin_line_num = 2


class MyThread(threading.Thread):
    def __init__(self, target, row_num, args=()):
        super().__init__()
        self.target = target
        self.row_num = row_num
        self.args = args  # args index, 0 is q，1 is q_type
        self._result = None

    def run(self):
        # print(f"now process with args {self.args}")
        self._result = self.target(*self.args)
        print(f"row {self.row_num} complete, q={self.args[0]}, q_type={self.args[1]},q_list={self._result}")

    def get_result(self):
        return self._result


def write_results(origin_excel, output_excel):
    global origin_line_num
    origin_workbook = openpyxl.load_workbook(origin_excel)
    output_workbook = openpyxl.load_workbook(output_excel)
    origin_sheet = origin_workbook['Sheet1']
    output_sheet = output_workbook['Sheet1']
    origin_col_count = origin_sheet.max_column
    origin_row_count = origin_sheet.max_row
    print("origin row count", origin_row_count)

    for i in range(2, origin_row_count + 1):
        q = origin_sheet.cell(row=i, column=1).value  # 店名
        try:
            q_list = results[i]
            if not q_list:
                continue
        except KeyError:
            continue
        if q not in q_list:
            print(f"q:{q} is not in q_list")
            q_list.insert(0, q)
        for j in range(0, len(q_list)):  # for each item in q_list
            for col in range(1, origin_col_count + 1):  # copy other info
                copy_cell = origin_sheet.cell(row=i, column=col).value
                output_sheet.cell(row=origin_line_num, column=col, value=copy_cell)
            output_sheet.cell(row=origin_line_num, column=origin_col_count + 1, value=q_list[j])
            origin_line_num += 1
    output_workbook.save(output_excel)


@retry(wait=wait_random(min=1, max=4), stop=stop_after_attempt(3))
def chat_completion_request(messages, model=GPT_MODEL):
    # print(f"{GREEN_BG} Info {END}, Generating ChatCompletion response")
    headers = {
        "Content-Type": "application/json",
        "Authorization": "Bearer " + openai.api_key,
    }
    json_data = {"model": model, "temperature": 0, "messages": messages}
    try:
        response = requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers=headers,
            json=json_data,
            timeout=30,
        )
        return response
    except Exception as e:
        print(f"{RED_BG} Warning {END}, Unable to generate ChatCompletion response")
        print(f"Exception: {type(e)} {e}")
        raise e


def get_q_list(q, q_type, sys_template):
    messages = []
    human_template = f"现在target_shop:{q}，q_type:{q_type}，以json格式对一组能够在电商平台上对target_shop进行搜索的搜索关键词q_list进行输出。q_list=？"
    input_template = sys_template + human_template
    messages.append({"role": "user", "content": input_template})
    try:
        chat_response = chat_completion_request(messages)
        # print(chat_response.json())
        assistant_message = chat_response.json()["choices"][0]["message"]
        content = assistant_message['content']
        q_list = format_q_list(content=content)
    except Exception as e:
        print(e)
        return []
    return q_list


def format_q_list(content):
    q_list = []
    q_list_str = ""
    try:
        input_string = re.sub(r'\s', '', content)  # Delete all whitespace
        input_string = input_string.replace('\\n', '')
        pattern = r'{"keywords":(.*?)}'
        match = re.search(pattern, input_string, re.DOTALL | re.UNICODE)
        if match:
            first_list = match.group(1)
            cleaned_content = re.sub(r'[\[\]\s]*', '', first_list)  # delete the extra [] and spaces
            q_list_str = cleaned_content
        else :
            print("first match failed")
            print(input_string)
            pattern = r'"keywords":\[(.*?)\]'  # process '{"keywords":[]'   ( without '}' )
            match = re.search(pattern, input_string, re.DOTALL | re.UNICODE)  
            if match:
                first_list = match.group(1)
                cleaned_content = re.sub(r'[\[\]\s]*', '', first_list)
                q_list_str = cleaned_content
                print("second match success")
            else:
                print("second match failed")

        q_list_str = q_list_str.replace('\"', '')
        q_list_str = q_list_str.replace('\'', '')
        q_list_str = q_list_str.replace('，', ',')
        q_list = list(q_list_str.split(","))

    except Exception as e:
        print(e)
    q_list = list(set(q_list))  # remove duplicate
    return q_list


def process(excel_name, sys_template):
    workbook = openpyxl.load_workbook(excel_name)
    sheet = workbook['Sheet1']
    row_count = sheet.max_row
    threads = []
    for row_num in range(2, row_count + 1):
        q = sheet.cell(row=row_num, column=shop_name_col).value
        q_type = sheet.cell(row=row_num, column=shop_type_col).value
        my_thread = MyThread(target=get_q_list, row_num=row_num, args=(q, q_type, sys_template))
        threads.append(my_thread)
    i = 0
    for t in threads:
        t.start()

        i += 1
        if i % 10 == 0 and i != 0:
            time.sleep(5)
        # gpt-3.5-turbo tokens per min (TPM) = 90000
        # https://platform.openai.com/account/limits
    for t in threads:
        t.join()
    for t in threads:
        q_list = t.get_result()
        # print(f"row: {t.row_num}, args: {t.args}, q_list: {q_list}")
        results[t.row_num] = q_list
        sheet.cell(row=t.row_num, column=empty_col, value=str(q_list))


def clean_excel(filename):
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active
    for row in worksheet.iter_rows():
        for cell in row:
            cell.value = None
    workbook.save(filename)


def main():
    global results
    # origin_excel = "shops_test.xlsx"
    origin_excel = "shops_v2.xlsx"
    target_excel = "gpt_gen_zero_shot.xlsx"
    # if target_excel does not exist, then create
    if not os.path.exists(target_excel):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        workbook.save(target_excel)
        print("create file "+target_excel)
    # else exists, clean it.
    else:
        print(f"clean file: {target_excel}")
        clean_excel(target_excel)

    print(f" origin excel:{origin_excel}, \n output_excel:{target_excel}.")
    process(excel_name=origin_excel,sys_template=sys_template_zero_shot)
    write_results(origin_excel, target_excel)
    return


if __name__ == "__main__":
    main()
