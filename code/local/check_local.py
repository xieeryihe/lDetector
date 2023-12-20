import requests
import json
import openpyxl
from tenacity import retry, RetryError, wait_random, stop_after_attempt

results = {}
shop_name_col = 1
shop_type_col = 5
keyword_col = 6
judgement_col = 7
origin_line_num = 2
target_line_num = 2

sys_template = '''
假设你是电商平台的搜索助理。该平台支持对各类商品和店铺的搜索，除搜索关键词外，返回结果时还考虑地理位置，优先返回附近的店铺和商品。
可以构建多种搜索关键词在电商平台上对店铺q进行搜索。搜索关键词的构造通常考虑店铺q的名称、店铺位置、店铺类型，店铺提供的服务等。
你的任务是已知店铺q和店铺类型q_type, 判断用户构建的搜索关键词keyword是否合理,合理的标准是能够完成对店铺q进行搜索，并且符合普通用户表达习惯。
judgement是你的判断结果，我将对你的判断结果进行解析，请严格按照要求输出，keyword合理输出“yes”,不合理输出“no”。
以下是几个合理的keyword，可供参考。
正例一 q="*****新疆菜（****广场店）"，q_type='新疆菜' ，keyword='** 新疆菜'，judgement='yes'
正例二 q="*******SPA（**店）"，q_type='按摩/足疗'， keyword='做SPA'，judgement='yes'
正例三 q="***艺术（****中心）"，q_type='美术'， keyword='画画 附近'，judgement='yes'
以下是几个不合理的keyword,可供参考。
反例一 q="***美甲美睫"，q_type='美甲' ，keyword='睫 ***美甲美'，judgement='no'
这个keyword不合理的原因是“睫”不是可以售卖的商品或服务，因为断句错误而不符合普通用户表达习惯
反例二 q="***美（***店）"，q_type='美甲'， keyword='美甲公众号'，judgement='no'
这个keyword不合理的原因是“美甲公众号”不是可以售卖的商品或服务，不符合普通用户表达习惯，并且不能对店铺q"***美（***店）"完成搜索
反例三 q="**造型（**路店）"，q_type='美发'， keyword='**附近剪头发'，judgement='no'
这个例子错误的原因是“**”是店名，不是地名，因此“**附近”不符合普通用户表达习惯，不是合理的搜索关键词。
反例四 q="***（*****汇店）"，q_type='火锅'， keyword='**汇'，judgement='no'
这个例子错误的原因是“**汇”是一个商场，范围远大于店铺q"***（*****汇店）"，不能完成对"***（*****汇店）"的搜索，因此不是合理的搜索关键词
'''

@retry(stop=stop_after_attempt(3))
def get_response(prompt):
    api_url = 'http://127.0.0.1:8000/'
    headers = {'Content-Type': 'application/json'}
    data = {
        'prompt': prompt,
    }
    data_json = json.dumps(data)
    value_str = ""
    try:
        response = requests.post(api_url, headers=headers, data=data_json, timeout=20)
        if response.status_code == 200:
            response_json = json.loads(response.text)
            print(response_json)
            value_str = str(response_json["response"])
        else:
            print("request failed, status code: ", response.status_code)
            print("error msg: ", response.text)
    except Exception as e:
        print(e)
        raise e
    return value_str

def format_judgement(response):
    if "yes" in response:
        return "yes"
    elif "not_sure" in response:
        return "not_sure"
    elif "no" in response:
        return "no"
    else:
        return "blank"
def process(origin_excel, target_excel):
    workbook = openpyxl.load_workbook(origin_excel)
    sheet = workbook['Sheet1']
    row_count = sheet.max_row
    for row in range(2, row_count + 1):
        process_row(row, sheet)
    workbook.save(target_excel)  

def process_row(row, sheet):
    q = sheet.cell(row=row, column=shop_name_col).value
    q_type = sheet.cell(row=row, column=shop_type_col).value
    keyword = sheet.cell(row=row, column=keyword_col).value
    human_template = f"现在q={q},q_type={q_type},keyword={keyword},judgement="
    input_template = sys_template + human_template
    try:
        response = get_response(input_template)
        judgement = format_judgement(response)
    except RetryError:
        print("Retry Error")
        judgement = "retry_error"
    except Exception:
        print("Unknow Error")
        judgement = "unknow_error"
    print(str(row) + human_template + judgement)
    sheet.cell(row=row, column=judgement_col, value=judgement)
    return False
    
def clean_excel(filename):
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active
    for row in worksheet.iter_rows():  
        for cell in row:  
            cell.value = None
    workbook.save(filename)

def main():

    origin_excel = "**.xlsx"
    target_excel = "**_check.xlsx"
    print(f" origin excel:{origin_excel}, \n output_excel:{target_excel}.")
    process(origin_excel=origin_excel, target_excel=target_excel)
    return

if __name__ == "__main__":
    main()
