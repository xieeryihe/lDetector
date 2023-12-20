import requests
import json
import openpyxl
import re
import os
from tenacity import retry, RetryError, wait_random, stop_after_attempt

results = {}
shop_name_col = 1
shop_type_col = 5
empty_col = 6
origin_line_num = 2

sys_template = '''
target_shop: *****新疆菜（****广场店），q_type:'新疆菜' 
Q: 以json格式对一组能够在电商平台上对target_shop进行搜索的搜索关键词q_list进行输出。q_list=？
A: 电商平台支持对各类商品和店铺的搜索，返回搜索结果时还考虑地理位置，优先返回附近的店铺和商品。
首先，target_shop的名称是"*****新疆菜（****广场店）",搜索店铺全称可以搜到该店，又从店铺名称出发，可搜索"**新疆"，"**"。
其次，从q_type看，target_shop应该是一家新疆菜餐厅，从店铺类型出发，可搜索"新疆菜"，"新疆餐厅" 。
再次，新疆菜餐厅通常提供大盘鸡、抓饭等菜品。从代表性菜品出发，可搜索"大盘鸡"，"抓饭" 。
最后，target_shop位于****广场，从地理位置出发，可搜索"新疆菜 **广场"，"** ****广场" ，"附近大盘鸡"。
因此：
q_list={{"keywords":[
"*****新疆菜（****广场店）",
"**新疆",
"**",
"新疆菜",
"新疆餐厅",
"大盘鸡",
"抓饭",
"新疆菜 **广场",
"** ****广场",
"附近大盘鸡"
]}}
target_shop: *******SPA（**店），q_type:'按摩/足疗'
Q: 以json格式对一组能够在电商平台上对target_shop进行搜索的搜索关键词q_list进行输出。q_list=？
A: 电商平台支持对各类商品和店铺的搜索，返回搜索结果时还考虑地理位置，优先返回附近的店铺和商品。
首先，target_shop的名称是"*******SPA（**店）",搜索店铺全称可以搜到该店，又从店铺名称出发，可搜索"****"，"日式SPA"。
其次，从q_type看，target_shop应该是一家按摩/足疗店，从店铺类型出发，可搜索"做SPA"，"做足疗"，"澡堂" 。
再次，按摩/足疗店通常提供泡澡、养生按摩等服务。从代表性服务出发，可搜索"泡澡"，"养生按摩" 。
最后，target_shop位于**，从地理位置出发，可搜索"泡澡 **"，"养生按摩 **" ，"附近泡澡"。
因此：
q_list={{"keywords":[
"*******SPA（**店）",
"****",
"日式SPA",
"做SPA",
"做足疗",
"澡堂",
"泡澡",
"养生按摩",
"泡澡 **",
"养生按摩 **",
"附近泡澡"
]}}
target_shop: ****（**路店），q_type:'超市'
Q: 以json格式对一组能够在电商平台上对target_shop进行搜索的搜索关键词q_list进行输出。q_list=？
A: 电商平台支持对各类商品和店铺的搜索，返回搜索结果时还考虑地理位置，优先返回附近的店铺和商品。
首先，target_shop的名称是"****（**路店）",搜索店铺全称可以搜到该店，又从店铺名称出发，可搜索"****"，"**"。
其次，从q_type看，target_shop应该是一家超市，从店铺类型出发，可搜索"大超市"，"**超市"，"****超市" ，"综合超市"，"逛超市"。
再次，超市通常水果生鲜等商品。从代表性商品出发，可搜索"超市买菜"。
最后，target_shop位于**路，从地理位置出发，可搜索"超市 **路"，"**** **路" ，"**超市 **路"，"附近****"。
因此：
q_list={{"keywords":[
"****（**路店）",
"****",
"**",
"大超市",
"**超市",
"****超市",
"综合超市",
"超市买菜",
"超市 **路",
"**** **路",
"**超市 **路",
"附近****"
]}}
target_shop: **影城（*****广场IMAX店）q_type:'电影院'
Q: 以json格式对一组能够在电商平台上对target_shop进行搜索的搜索关键词q_list进行输出。q_list=？
A: 电商平台支持对各类商品和店铺的搜索，返回搜索结果时还考虑地理位置，优先返回附近的店铺和商品。
首先，target_shop的名称是"**影城（*****广场IMAX店）",搜索店铺全称可以搜到该店，又从店铺名称出发，可搜索"**影城"，"影城"。
其次，从q_type看，target_shop应该是一家电影院，从店铺类型出发，可搜索"电影院"，"影院"。
再次，电影院通常提供观影服务。从代表性服务出发，可搜索"看电影"，"IMAX电影"，"3D电影"，"电影新片"。
最后，target_shop位于*****广场，从地理位置出发，可搜索"电影院 ***"，"**广场的电影院" ，"**影城 ***"，"附近的电影院"。
因此：
q_list={{"keywords":[
"**影城（*****广场IMAX店）",
"**影城",
"影城",
"电影院",
"影院",
"看电影",
"IMAX电影",
"3D电影",
"电影新片",
"影院 ***",
"**广场的电影院",
"**影城 ***",
"附近的电影院"
]}}
target_shop: **汽车工作室，q_type:'汽车维修'
Q: 以json格式对一组能够在电商平台上对target_shop进行搜索的搜索关键词q_list进行输出。q_list=？
A: 电商平台支持对各类商品和店铺的搜索，返回搜索结果时还考虑地理位置，优先返回附近的店铺和商品。
首先，target_shop的名称是"**汽车工作室",搜索店铺全称可以搜到该店，又从店铺名称出发，可搜索"**汽车"，"汽车工作室"。
其次，从q_type看，target_shop应该是一家汽车维修店，从店铺类型出发，可搜索"修车铺"，"修车店"，"修汽车"。
再次，汽车维修店提供换轮胎，汽车喷漆，汽车改装等服务。从代表性服务出发，可搜索"汽车换轮胎"，"汽车喷漆"，"汽车改装"。
最后，target_shop未提供地理位置信息，从地理位置出发，可搜索"修车店 附近"，"附近修汽车的地方" 。
因此：
q_list={{"keywords":[
"**汽车工作室",
"**汽车",
"汽车工作室",
"修车铺",
"修车店",
"修汽车",
"汽车换轮胎",
"汽车喷漆",
"汽车改装",
"修车店 附近",
"附近修汽车的地方"
]}}
保证生成的 keywords 列表最多包含 7 个条目。
'''


def write_results(origin_excel, output_excel):
    global origin_line_num
    origin_workbook = openpyxl.load_workbook(origin_excel)
    output_workbook = openpyxl.load_workbook(output_excel)
    origin_sheet = origin_workbook['Sheet1']
    output_sheet = output_workbook['Sheet1']
    origin_col_count = origin_sheet.max_column
    origin_row_count = origin_sheet.max_row
    print("origin row count", origin_row_count)

    for i in range(2, origin_row_count + 2):
        q = origin_sheet.cell(row=i, column=1).value
        try:
            q_list = results[i]
            if not q_list:
                continue
        except KeyError:
            continue
        if q not in q_list:
            print(f"q:{q} is not in q_list")
            q_list.insert(0, q)
        for j in range(0, len(q_list)):  # for each item in q_list
            for col in range(1, origin_col_count + 1):  # copy other info
                copy_cell = origin_sheet.cell(row=i, column=col).value
                output_sheet.cell(row=origin_line_num, column=col, value=copy_cell)
            output_sheet.cell(row=origin_line_num, column=origin_col_count + 1, value=q_list[j])
            origin_line_num += 1
    output_workbook.save(output_excel)

@retry(stop=stop_after_attempt(3))
def get_q_list(prompt):
    api_url = 'http://127.0.0.1:8000/'
    headers = {'Content-Type': 'application/json'}
    data = {
        'prompt': prompt,
    }

    data_json = json.dumps(data)
    try:
        response = requests.post(api_url, headers=headers, data=data_json, timeout=20)
    except requests.exceptions.ReadTimeout:
        print("timeout")
        return []

    q_list = None
    if response.status_code == 200:
        q_list_str = ""
        try:
            response_json = json.loads(response.text)
            value_str = str(response_json["response"])
            # print(value_str)
            
            # format response
            input_string = re.sub(r'\s', '', value_str)
            input_string = input_string.replace('\\n', '')  # delete \n 
            pattern = r'{"keywords":(.*?)}'
            match = re.search(pattern, input_string, re.DOTALL | re.UNICODE)  
            if match:
                first_list = match.group(1)
                cleaned_content = re.sub(r'[\[\]\s]*', '', first_list)  # delete the extra [] and spaces
                q_list_str = cleaned_content
            else :
                print("first match failed")
                print(input_string)
                pattern = r'"keywords":\[(.*?)\]'  # process '{"keywords":[]'   ( without '}' )
                match = re.search(pattern, input_string, re.DOTALL | re.UNICODE)  
                if match:
                    first_list = match.group(1)
                    cleaned_content = re.sub(r'[\[\]\s]*', '', first_list)
                    q_list_str = cleaned_content
                    print("second match success")
                else:
                    print("second match failed")

            q_list_str = q_list_str.replace('\"', '')
            q_list_str = q_list_str.replace('\'', '')
            q_list_str = q_list_str.replace('，', ',')
            q_list = list(q_list_str.split(","))

        except Exception as e:
            print(e)
    else:
        print("request failed, status code: ", response.status_code)
        print("error msg: ", response.text)
    if q_list is None:
        q_list = []
    return q_list


def process(excel_name):
    workbook = openpyxl.load_workbook(excel_name)
    sheet = workbook['Sheet1']
    row_count = sheet.max_row
    for row in range(2, row_count + 1):
        r = process_row(row, sheet)
        if not r:  # Excel file may has empty lines at the end of it.
            return

def process_row(row, sheet):
    q = sheet.cell(row=row, column=shop_name_col).value
    if q is None:
        print(f"row: {row} is empty")
        return False
    q_type = sheet.cell(row=row, column=shop_type_col).value
    human_template = f"现在target_shop:{q}，q_type:{q_type}，以json格式对一组能够在电商平台上对target_shop进行搜索的搜索关键词q_list进行输出。q_list=？"
    prompt = sys_template + human_template
    print(f"now process row {row}, q = {q}")
    q_list = []
    try:
        q_list = get_q_list(prompt)
    except Exception as e:
        print(e)
    q_list = list(set(q_list))  # remove duplicate 
    q_list = [str for str in q_list if str != ""]
    results[row] = q_list
    # print(q_list)
    
    return True
    
def clean_excel(filename):
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active
    for row in worksheet.iter_rows():  
        for cell in row:  
            cell.value = None
    workbook.save(filename)

def main():
    origin_excel = "shops.xlsx"
    target_excel = "answer.xlsx"
    # if target_excel does not exist, then create
    if not os.path.exists(target_excel):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        workbook.save(target_excel)
        print("create file "+target_excel)
    # else exists, clean it.
    else:
        print(f"clean file: {target_excel}")
        clean_excel(target_excel)

    print(f" origin excel:{origin_excel}, \n output_excel:{target_excel}.")
    process(origin_excel)
    write_results(origin_excel, target_excel)
    return


if __name__ == "__main__":
    main()
